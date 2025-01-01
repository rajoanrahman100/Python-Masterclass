import openpyxl as xl

from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx") # Load the workbook

sheet = wb["Sheet1"] # Load the sheet in the workbook

# cell = sheet["a1"] # Load the cell in the sheet

# print(cell.value) # Print the value of the cell

# print(sheet.max_row) # Print the maximum row in the sheet

# print(sheet.max_column) # Print the maximum column in the sheet

for row in range(2 , sheet.max_row+1): # Loop through the rows in the sheet
    
    cell=sheet.cell(row,3) # Load the cell in the sheet
    corrected_price=cell.value * 0.9
    corrected_price_cell=sheet.cell(row,4) # Load the cell in the sheet
    corrected_price_cell.value=corrected_price # Assign the value to the cell
    
wb.save("transactions2.xlsx") # Save the workbook

values=Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)

chart=BarChart() # Create the chart

chart.add_data(values) # Add the data to the chart

sheet.add_chart(chart,"e2") # Add the chart to the sheet